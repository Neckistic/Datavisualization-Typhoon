import openpyxl, json

wb = openpyxl.load_workbook('TC_Impact_Data_HKO.xlsx', data_only=True)

def sf(v):
    if v is None or v == 'NR' or v == 'NA':
        return 0
    try:
        return float(v)
    except:
        return 0

def si(v):
    if v is None or v == 'NR' or v == 'NA':
        return 0
    try:
        return int(v)
    except:
        return 0

def sf_none(v):
    if v is None or v == 'NA' or v == 'NR':
        return None
    try:
        return float(v)
    except:
        return None

# ===== Signal Data =====
ws_sig = wb['Signal Data']
signal_list = []
for row in ws_sig.iter_rows(min_row=7, max_row=ws_sig.max_row, values_only=True):
    if row[0] is None:
        continue
    signal_list.append({
        'year': int(row[0]), 'name': str(row[1]), 'signal': int(row[4]),
        'intensity': str(row[9]) if row[9] else None,
        'pressure': sf_none(row[10]),
        'distance': sf_none(row[11]),
        'bearing': str(row[12]) if row[12] else None
    })

# ===== Other Met Data (surge) =====
ws_met = wb['Other Met Data']
met_lookup = {}
for row in ws_met.iter_rows(min_row=6, max_row=ws_met.max_row, values_only=True):
    if row[0] is None:
        continue
    key = str(int(row[0])) + '_' + str(row[1])
    met_lookup[key] = {'surge': sf_none(row[7])}

# ===== Casualty =====
ws_cas = wb['Casualty and Vessel Damage']
cas_lookup = {}
for row in ws_cas.iter_rows(min_row=6, max_row=ws_cas.max_row, values_only=True):
    if row[0] is None:
        continue
    key = str(int(row[0])) + '_' + str(row[1])
    cas_lookup[key] = {
        'dead': si(row[5]), 'missing': si(row[6]),
        'injured': si(row[7]), 'affected': si(row[8])
    }

# ===== Damage (Monetary) =====
# Formula: subtotal = sum(agriculture..others), col 5-10
# insurance_est = (col12 + col13) / col14  (property+EC / market_share)
# total = subtotal + insurance_est
ws_dmg = wb['Damage (Monetary)']
damage_lookup = {}
for row in ws_dmg.iter_rows(min_row=6, max_row=ws_dmg.max_row, values_only=True):
    if row[0] is None:
        continue
    key = str(int(row[0])) + '_' + str(row[1])
    agriculture = sf(row[5])
    publicWorks = sf(row[6])
    utilities = sf(row[7])
    privateProperty = sf(row[8])
    industry = sf(row[9])
    others = sf(row[10])
    subtotal_calc = agriculture + publicWorks + utilities + privateProperty + industry + others
    
    # Insurance claims: (col12 + col13) / col14
    prop_ins = sf(row[12])
    ec_ins = sf(row[13])
    market_share = sf(row[14])
    insurance_est = 0
    if market_share > 0:
        insurance_est = (prop_ins + ec_ins) / market_share
    
    # Use cached total if available, otherwise compute
    cached_total = sf(row[16])
    if cached_total > 0:
        total_loss = cached_total
    else:
        total_loss = subtotal_calc + insurance_est
    
    # Use cached subtotal if available
    cached_sub = sf(row[11])
    subtotal = cached_sub if cached_sub > 0 else subtotal_calc
    
    damage_lookup[key] = {
        'totalLoss': round(total_loss, 2),
        'subtotal': round(subtotal, 2)
    }

# ===== Wind Data =====
ws_wind = wb['Wind']
row2 = list(ws_wind.iter_rows(min_row=2, max_row=2, values_only=True))[0]
wind_stations = []
for i in range(7, len(row2), 2):
    if row2[i] is not None:
        sname = str(row2[i]).strip().replace('\n', ' ')
        wind_stations.append({'name': sname, 'gustCol': i, 'meanCol': i + 1})

wind_lookup = {}
for row in ws_wind.iter_rows(min_row=5, max_row=ws_wind.max_row, values_only=True):
    if row[2] is None:
        continue
    key = str(int(row[2])) + '_' + str(row[3])
    stations = {}
    for st in wind_stations:
        gust = row[st['gustCol']] if st['gustCol'] < len(row) else None
        mean_w = row[st['meanCol']] if st['meanCol'] < len(row) else None
        if gust == 'NA' or gust == 'NR':
            gust = None
        if mean_w == 'NA' or mean_w == 'NR':
            mean_w = None
        if gust is not None or mean_w is not None:
            stations[st['name']] = {'gust': sf_none(gust), 'mean': sf_none(mean_w)}
    wind_lookup[key] = stations

# ===== Rainfall Data =====
ws_rain = wb['Rainfall']
row2r = list(ws_rain.iter_rows(min_row=2, max_row=2, values_only=True))[0]
rain_stations = []
for i in range(5, len(row2r)):
    if row2r[i] is not None:
        sname = str(row2r[i]).strip().replace('\n', ' ')
        rain_stations.append({'name': sname, 'col': i})

rain_lookup = {}
for row in ws_rain.iter_rows(min_row=4, max_row=ws_rain.max_row, values_only=True):
    if row[0] is None:
        continue
    key = str(int(row[0])) + '_' + str(row[1])
    stations = {}
    for st in rain_stations:
        val = row[st['col']] if st['col'] < len(row) else None
        if val == 'NA' or val == 'NR' or val is None:
            continue
        try:
            stations[st['name']] = float(val)
        except:
            pass
    rain_lookup[key] = stations

# ===== Merge all =====
merged = []
for s in signal_list:
    key = str(s['year']) + '_' + s['name']
    m = met_lookup.get(key, {})
    c = cas_lookup.get(key, {})
    d = damage_lookup.get(key, {})
    w = wind_lookup.get(key, {})
    r = rain_lookup.get(key, {})

    # Get max gust across all stations
    gusts = [v['gust'] for v in w.values() if v.get('gust') is not None]
    maxGust = max(gusts) if gusts else 0

    # Get HKO HQ rainfall (last rain station) or max across stations
    hko_rain = r.get('HKO Headquarters', 0) or 0
    all_rain = [v for v in r.values() if v > 0]
    maxRain = max(all_rain) if all_rain else 0

    merged.append({
        'year': s['year'], 'name': s['name'], 'signal': s['signal'],
        'intensity': s.get('intensity'), 'distance': s.get('distance'),
        'bearing': s.get('bearing'), 'pressure': s.get('pressure'),
        'surge': m.get('surge') or 0,
        'maxGust': maxGust, 'maxRain': maxRain, 'hkoRain': hko_rain,
        'dead': c.get('dead', 0), 'missing': c.get('missing', 0),
        'injured': c.get('injured', 0), 'affected': c.get('affected', 0),
        'loss': d.get('totalLoss', 0),
        'windStations': w if w else None,
        'rainStations': r if r else None
    })

print(f'Merged records: {len(merged)}')
nonzero_loss = sum(1 for r in merged if r['loss'] > 0)
print(f'Non-zero loss: {nonzero_loss}')

# Show high-loss entries
for r in merged:
    if r['loss'] > 100:
        print(f"  {r['year']} {r['name']}: signal={r['signal']}, loss={r['loss']:.1f}M, affected={r['affected']}")

with open('tc_data.json', 'w', encoding='utf-8') as f:
    json.dump(merged, f, ensure_ascii=False)
print('\nExported tc_data.json')

# Also export station data separately
station_wind = [{'name': s['name']} for s in wind_stations]
station_rain = [{'name': s['name']} for s in rain_stations]
print(f'\nWind stations: {len(station_wind)}')
print(f'Rain stations: {len(station_rain)}')
